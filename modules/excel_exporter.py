# modules/excel_exporter.py
# ─────────────────────────────────────────────────────────────────────────────
# Excel export for AssistantAgent render_table results.
#
# Public API:
#   export_table_to_excel(table_data: dict, question: str, project_name: str) -> bytes
#
# table_data keys (match render_table tool schema):
#   title       str           — sheet/chart title
#   columns     list[str]     — column headers
#   rows        list[list]    — data rows (each row = list of values)
#   chart_type  str|None      — "bar" | "pie" | "line" | None
#   chart_x_col str|None      — column name for X axis (bar/line) or labels (pie)
#   chart_y_cols list[str]|None — column names for Y axis series
# ─────────────────────────────────────────────────────────────────────────────

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter


# ── Brand colours (aligned with Process2Diagram IBM Plex theme) ───────────────
_NAVY       = "0F172A"
_ZEBRA      = "F1F5F9"
_WHITE      = "FFFFFF"
_BORDER_CLR = "CBD5E1"

_HEADER_FONT = Font(name="Calibri", bold=True, color=_WHITE, size=11)
_BODY_FONT   = Font(name="Calibri", size=10)
_TITLE_FONT  = Font(name="Calibri", bold=True, size=13, color=_NAVY)

_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_ZEBRA_FILL  = PatternFill("solid", fgColor=_ZEBRA)

_THIN_BORDER = Border(
    left=Side(style="thin", color=_BORDER_CLR),
    right=Side(style="thin", color=_BORDER_CLR),
    top=Side(style="thin", color=_BORDER_CLR),
    bottom=Side(style="thin", color=_BORDER_CLR),
)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Public entry point ────────────────────────────────────────────────────────

def export_table_to_excel(
    table_data: dict[str, Any],
    question: str = "",
    project_name: str = "",
) -> bytes:
    """
    Build a .xlsx file in memory and return raw bytes suitable for
    st.download_button(data=...).
    """
    wb = Workbook()

    ws_data = wb.active
    ws_data.title = _safe_sheet_name(table_data.get("title", "Dados"))

    _write_data_sheet(
        ws=ws_data,
        title=table_data.get("title", ""),
        columns=table_data.get("columns", []),
        rows=table_data.get("rows", []),
    )

    chart_type   = (table_data.get("chart_type") or "").lower()
    chart_x_col  = table_data.get("chart_x_col")
    chart_y_cols = table_data.get("chart_y_cols") or []

    if chart_type in ("bar", "pie", "line") and chart_x_col and chart_y_cols:
        ws_chart = wb.create_sheet(title="Grafico")
        _write_chart_sheet(
            ws_chart=ws_chart,
            ws_data=ws_data,
            table_data=table_data,
            chart_type=chart_type,
            chart_x_col=chart_x_col,
            chart_y_cols=chart_y_cols,
        )

    ws_meta = wb.create_sheet(title="Metadados")
    _write_metadata_sheet(ws_meta, table_data, question, project_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Data sheet ────────────────────────────────────────────────────────────────

def _write_data_sheet(ws, title: str, columns: list[str], rows: list[list]) -> None:
    if not columns:
        return

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title or "Dados")
    title_cell.font      = _TITLE_FONT
    title_cell.alignment = _CENTER
    title_cell.fill      = PatternFill("solid", fgColor="E2E8F0")
    ws.row_dimensions[1].height = 28

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER
    ws.row_dimensions[2].height = 22

    for row_idx, row in enumerate(rows, start=3):
        fill = _ZEBRA_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = _BODY_FONT
            cell.alignment = _LEFT
            cell.border    = _THIN_BORDER
            if fill:
                cell.fill = fill

    for col_idx, col_name in enumerate(columns, start=1):
        col_values = [str(col_name)] + [
            str(r[col_idx - 1]) for r in rows if col_idx - 1 < len(r)
        ]
        max_len = min(max((len(v) for v in col_values), default=10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}{len(rows) + 2}"

    ws.freeze_panes = "A3"


# ── Chart sheet ───────────────────────────────────────────────────────────────

def _write_chart_sheet(
    ws_chart, ws_data,
    table_data: dict,
    chart_type: str,
    chart_x_col: str,
    chart_y_cols: list[str],
) -> None:
    columns   = table_data.get("columns", [])
    rows      = table_data.get("rows", [])
    title_str = table_data.get("title", "")

    if not columns or not rows:
        return

    try:
        x_col_idx = columns.index(chart_x_col) + 1
    except ValueError:
        x_col_idx = 1

    y_col_indices = []
    for yc in chart_y_cols:
        try:
            y_col_indices.append(columns.index(yc) + 1)
        except ValueError:
            pass

    if not y_col_indices:
        return

    data_start_row = 3
    data_end_row   = data_start_row + len(rows) - 1

    if chart_type == "bar":
        chart = BarChart()
        chart.type     = "col"
        chart.grouping = "clustered"
        chart.overlap  = 0
    elif chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        return

    chart.title  = title_str
    chart.style  = 10
    chart.width  = 22
    chart.height = 14

    cats = Reference(ws_data, min_col=x_col_idx, min_row=data_start_row, max_row=data_end_row)

    for y_idx in y_col_indices:
        data_ref = Reference(ws_data, min_col=y_idx, min_row=data_start_row - 1, max_row=data_end_row)
        chart.add_data(data_ref, titles_from_data=True)

    chart.set_categories(cats)

    if chart_type == "bar":
        chart.shape = 4
    elif chart_type == "pie" and chart.series:
        chart.series[0].explosion = 10

    ws_chart.add_chart(chart, "B2")
    ws_chart["A1"].value     = f"Grafico - {title_str}"
    ws_chart["A1"].font      = _TITLE_FONT
    ws_chart["A1"].alignment = _LEFT


# ── Metadata sheet ────────────────────────────────────────────────────────────

def _write_metadata_sheet(ws, table_data: dict, question: str, project_name: str) -> None:
    meta_rows = [
        ("Projeto",          project_name),
        ("Exportado em",     datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Titulo da tabela", table_data.get("title", "")),
        ("Pergunta original", question),
        ("Colunas",          ", ".join(table_data.get("columns", []))),
        ("Total de linhas",  len(table_data.get("rows", []))),
        ("Tipo de grafico",  table_data.get("chart_type") or "nenhum"),
        ("Gerado por",       "Process2Diagram - AgentAssistant"),
    ]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    for r_idx, (label, value) in enumerate(meta_rows, start=1):
        lc = ws.cell(row=r_idx, column=1, value=label)
        vc = ws.cell(row=r_idx, column=2, value=str(value))
        lc.font      = Font(name="Calibri", bold=True, size=10, color=_NAVY)
        vc.font      = _BODY_FONT
        lc.alignment = _LEFT
        vc.alignment = _LEFT


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_sheet_name(name: str) -> str:
    """Truncate to 31 chars and strip characters illegal in Excel sheet names."""
    for ch in r'\/*?:[]':
        name = name.replace(ch, "")
    return name[:31] or "Dados"
